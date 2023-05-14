import requests  # HTTPリクエストを送るためのライブラリ
import json  # JSON形式のデータを扱うためのライブラリ
from openpyxl import Workbook  # Excelファイルを作成・操作するためのライブラリ
from openpyxl.styles import PatternFill, Protection  # Excelのセルの塗りつぶしと保護設定をするためのライブラリ
from openpyxl.utils import get_column_letter  # 数値をExcelの列のアルファベット（例：1 -> 'A'）に変換するための関数
from datetime import datetime, timedelta  # 日時と時間差を扱うためのライブラリ

# 設定ファイルを読み込みます。
with open("config.json") as f:
    config = json.load(f)

# Mattermost のエンドポイントとアクセストークンを設定します。
url = config["url"]
token = config["token"]

# ユーザー一覧を取得します。
user_list_url = f"{url}/api/v4/users"
headers = {"Authorization": f"Bearer {token}"}
response = requests.get(user_list_url, headers=headers)
users = response.json()

# ユーザーIDとユーザーネームをマッピングする辞書を作成します。
user_dict = {user["id"]: user["username"] for user in users}

# Excelファイルを作成し、デフォルトのシートを削除します。
wb = Workbook()
del wb["Sheet"]

# チームごとに各チャンネルのユーザー参加情報を保持するための辞書を作成します。
team_channel_data = {}

# 現在時刻を取得して、状況表示用に使用します。
start_time = datetime.now()

# 各ユーザーに対して処理を行います。
for user_id, username in user_dict.items():
    # ユーザーが所属するチーム一覧を取得します。
    teams_url = f"{url}/api/v4/users/{user_id}/teams"
    response = requests.get(teams_url, headers=headers)
    teams = response.json()

    # 各チームについて、ユーザーが所属するチャンネルを取得します。
    for team in teams:
        # チームの ID を取得します。
        team_id = team["id"]

        # ユーザーが所属するチャンネル一覧を取得します。
        channels_url = f"{url}/api/v4/users/{user_id}/teams/{team_id}/channels"
        response = requests.get(channels_url, headers=headers)
        channels = response.json()

        # 各チャンネルについて処理を行います。
        for channel in channels:
            # チャンネルの ID を取得します。
            channel_id = channel["id"]

            # チャンネル名が空の場合は次のチャンネルへ移ります。
            if not channel["display_name"]:
                continue

            # チャンネルのメンバー一覧を取得します。
            channel_members_url = f"{url}/api/v4/channels/{channel_id}/members"
            response = requests.get(channel_members_url, headers=headers)
            channel_members = [member["user_id"] for member in response.json()]

            # チャンネルの参加状況情報を辞書に追加します。
            team_channel_data.setdefault(team["display_name"], {}).setdefault(
                channel["display_name"], {}
            )[user_id] = ("〇" if user_id in channel_members else "")

        # 1 分ごとに実行状況を表示します。
        elapsed_time = datetime.now() - start_time
        print(
            f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} (経過時間: {str(elapsed_time).split('.')[0]}) {team['display_name']} チームのチャンネルの一覧を取得しています..."
        )

# 各チームのデータをExcelに書き出します。
for team, channel_data in team_channel_data.items():
    # 新しいシートを作成し、チーム名を設定します。
    ws = wb.create_sheet(title=team)

    # 列名と詳細情報を作成します。
    headers = ["ユーザー名"]
    header_details = []  # 各列の詳細情報を保存するリスト
    for channel in channel_data.keys():
        if not channel:  # チャンネル名が存在しない場合はスキップ
            continue
        # アーカイブされたチャンネルを無視
        if channel_data[channel].get("is_archived", False):
            continue
        header = f"{channel}"
        headers.append(header)
        header = f"(指示)"
        headers.append(header)
        header_details.append(channel_data[channel])  # 各列の詳細情報を保存

    # 列名を書き込みます。
    ws.append(headers)

    # 各ユーザーの参加状況を書き出します。
    for user_id, username in user_dict.items():
        row = [username]

        should_write_row = False
        for channel in channel_data.keys():
            current_status = channel_data[channel].get(user_id, "")
            next_status = channel_data[channel].get(user_id, "")
            row.append(current_status)
            row.append(next_status)

            # ユーザーが何らかのチャンネルで現在のステータスを持っている場合、行を書き込みます。
            if current_status:
                should_write_row = True

        # ユーザーがチーム内のどのチャンネルにも参加していない場合は、その行を書き込まない
        if should_write_row:
            ws.append(row)

    # EXCELの装飾を行います。
    # 背景色を設定します。
    header_fill = PatternFill(
        start_color="FFFF99", end_color="FFFF99", fill_type="solid"
    )
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    dark_grey_fill = PatternFill(
        start_color="808080", end_color="808080", fill_type="solid"
    )  # プライベートチャンネル用の背景色を追加
    alternate_fill = PatternFill(
        start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
    )

    # チャンネル名の列の背景色を変更します。
    for i, header_detail in enumerate(header_details, start=1):  # インデックスを1から開始
        for cell in ws[get_column_letter(i * 2)]:
            if header_detail.get("type", "") == "private":  # チャンネルがプライベートの場合
                cell.fill = dark_grey_fill
            else:
                cell.fill = grey_fill
            cell.protection = Protection(locked=True)

    # 1行おきに背景色を設定します。
    for row in ws.iter_rows(min_row=2):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = alternate_fill

    # ユーザー名の列とチャンネル名列の幅を2倍にします。
    ws.column_dimensions["A"].width = ws.column_dimensions["A"].width * 2
    for i in range(2, len(headers) * 2, 2):
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = (
            ws.column_dimensions[column_letter].width * 2
        )

    # チャンネル名の列を保護します。
    for i in range(2, len(headers) * 2, 2):
        for cell in ws[get_column_letter(i)]:
            cell.fill = grey_fill
            cell.protection = Protection(locked=True)

    # "(指示)"の列は変更可能な列として設定します。
    for i in range(3, len(headers) * 2 + 1, 2):
        for cell in ws[get_column_letter(i)]:
            cell.protection = Protection(locked=False)

    # シート全体の保護を有効にします。
    ws.protection.sheet = True

    # 1行目と1列目を固定します。
    ws.freeze_panes = "B2"

# Excelファイルを保存します。
try:
    wb.save(config["excel_dir"] + config["excel_file"])
    print(f"エクセルファイルを保存しました")
except Exception as e:
    print(f"エクセルファイルの保存中にエラーが発生しました: {e}")
