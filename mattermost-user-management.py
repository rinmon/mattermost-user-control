import os
import json
import requests
import sys
from openpyxl import load_workbook
from urllib.parse import quote

# 設定ファイルを読み込みます。
try:
    with open("config.json") as f:
        config = json.load(f)
except (FileNotFoundError, json.JSONDecodeError) as e:
    print(f"設定ファイルの読み込みに失敗しました: {e}")
    sys.exit(1)

# Mattermost のエンドポイントとアクセストークンを設定します。
url = config["url"]
token = config["token"]
headers = {"Authorization": f"Bearer {token}"}

# Excelファイルを読み込みます。
workbook_path = os.path.join(config["excel_dir"], config["excel_file"])
wb = load_workbook(workbook_path)

# 出力ファイルを作成します。
output_file_path1 = os.path.join(config["excel_dir"], "output_all.txt")
output_file_path2 = os.path.join(config["excel_dir"], "output_changes_only.txt")


def get_team_id_by_name(team_name):
    response = requests.post(
        f"{url}/api/v4/teams/search",
        headers=headers,
        json={"term": team_name},
    )

    response.raise_for_status()

    try:
        teams_data = response.json()
        if not teams_data:
            print(f"'{team_name}'という名前のチームが見つかりませんでした")
            return None
    except ValueError:
        print(f"チーム {team_name} のJSONデータ解析エラー: {response.text}")
        return None

    return teams_data[0]["id"]


def create_channel_mapping(team_id):
    # パブリックチャンネルを取得
    public_channels_response = requests.get(
        f"{url}/api/v4/teams/{team_id}/channels", headers=headers
    )
    public_channels_response.raise_for_status()

    # パブリックチャンネルを取得
    private_channels_response = requests.get(
        f"{url}/api/v4/users/me/teams/{team_id}/channels", headers=headers
    )
    private_channels_response.raise_for_status()

    try:
        public_channels_data = public_channels_response.json()
        private_channels_data = private_channels_response.json()
        channels_data = public_channels_data + private_channels_data
    except ValueError:
        print(
            f"チーム {team_id} のJSONデータ解析エラー: {public_channels_response.text} {private_channels_response.text}"
        )
        return {}

    if not isinstance(channels_data, list):
        print(f"channels_dataの形式が予期せぬものです: {channels_data}")
        return {}

    return {channel["display_name"]: channel["id"] for channel in channels_data}


# Output API responses to a text file
api_responses_file_path = os.path.join(config["excel_dir"], "api_responses.txt")

with open(api_responses_file_path, "w", encoding="utf-8") as api_responses_file:
    for sheet_name in wb.sheetnames:
        team_id = get_team_id_by_name(sheet_name)
        api_responses_file.write(f"Team: {sheet_name} - Team ID: {team_id}\n")

        channel_mapping = create_channel_mapping(team_id)
        for channel_name, channel_id in channel_mapping.items():
            api_responses_file.write(
                f"Channel: {channel_name} - Channel ID: {channel_id}\n"
            )
        api_responses_file.write("\n")


with open(output_file_path1, "w", encoding="utf-8") as output_file1, open(
    output_file_path2, "w", encoding="utf-8"
) as output_file2:
    # 全シートを読み込みます。
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # シート名＝チーム名の一覧を出力します。
        team_title = f"チーム名: {sheet_name}"
        output_file1.write(f"{team_title}\n")
        output_file2.write(f"{team_title}\n")

        # チーム名ごとに横線を引く
        separator_line = "=" * (len(team_title) + 2)
        output_file1.write(f"{separator_line}\n\n")
        output_file2.write(f"{separator_line}\n\n")

        # チームIDとチャンネルマッピングを取得
        team_id = get_team_id_by_name(sheet_name)
        channel_mapping = create_channel_mapping(team_id)

        # 各シートの各行について処理を行います。
        for row in ws.iter_rows(min_row=2):
            username = row[0].value

            # ユーザー名からユーザーIDを取得します。
            user_response = requests.get(
                f"{url}/api/v4/users/username/{username}", headers=headers
            )
            user_data = user_response.json()
            user_id = user_data["id"]

            # ユーザー毎に各チャンネルへの参加状況と指示内容を出力します。
            output_file1.write(f"ユーザー名: {username}\n")
            has_changes = False
            changes_output = []

            for col_idx in range(1, len(row), 2):
                channel_name = ws.cell(row=1, column=col_idx + 1).value
                current_status = row[col_idx].value
                next_status = row[col_idx + 1].value
                api_result = ""

                # 現在の参加状況と指示が同じかどうかを判断し、出力する内容を決定します。
                if current_status == next_status:
                    status_change = "変更なし"
                else:
                    channel_id = channel_mapping[channel_name]

                    if next_status == "〇":
                        status_change = "招待"
                        # Mattermost API でユーザーを該当のチームのチャンネルにJOINさせる
                        join_response = requests.post(
                            f"{url}/api/v4/channels/{channel_id}/members",
                            headers=headers,
                            json={"user_id": user_id},
                        )
                        api_result = (
                            f"API結果: {join_response.status_code} {join_response.text}"
                        )

                    elif current_status == "〇":
                        status_change = "退会"
                        # Mattermost API でユーザーを該当のチームのチャンネルからleaveする
                        leave_response = requests.delete(
                            f"{url}/api/v4/channels/{channel_id}/members/{user_id}",
                            headers=headers,
                        )
                        api_result = (
                            f"API結果: {leave_response.status_code} {leave_response.text}"
                        )

                    else:
                        status_change = f"{current_status} -> {next_status}"

                    has_changes = True

                # 参加状況と指示内容をわかりやすく出力します。
                change_line = f"チャンネル名: {channel_name} | 現在の参加状況: {'参加' if current_status == '〇' else '不参加'} | 指示: {status_change} | {api_result}\n"
                output_file1.write(change_line)

                if status_change != "変更なし":
                    changes_output.append(change_line)

            # 変更がある場合はサマリーファイルに出力します。
            if has_changes:
                output_file2.write(f"ユーザー名: {username}\n")
                output_file2.writelines(changes_output)
                output_file2.write("\n")

            output_file1.write("\n")

print(f"テキストファイルに処理内容を出力しました: {output_file_path1}")
print(f"変更があったユーザーのみを出力したテキストファイルを作成しました: {output_file_path2}")
